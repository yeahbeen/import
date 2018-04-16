'''
作用：
    把excel的用例转化成能导入到testlink的xml
使用方法：
    1、把用例Excel文件和该脚本放到同一个目录
    2、命令行进入该目录，执行python import_case.py
    3、执行结束后，就会生成能导入到testlink的xml，每个Excel对应一个文件夹，Excel的每个sheet对应一个xml。
使用条件：
    1、安装Python3，安装openpyxl包
    2、Excel要符合一定的格式：
        1）、第一列留空，从第二列(B列)开始，依次是：用例编号、模块名称、用例名称、前置条件、操作步骤、期望输出    
        2）、正式的用例从11行开始，11行以前的内容忽略
        3）、每个用例都要有编号，编号为空将会结束解析
        4）、Sheet1会被忽略，后面每个sheet作为一个大模块

'''
import os,sys
from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange

dir = __file__+"\\..\\"

filename = []
for file in os.listdir(dir):
    # print(file)
    # print(os.path.splitext(file)[1])
    if os.path.splitext(file)[1] == ".xlsx": 
        filename.append(file)
        
if filename == []:
    print("没有Excel文件")
    sys.exit()
    
    
        
def ismerged(cell):
    for r in sheet_ranges.merged_cells.ranges:
        if r.issuperset(CellRange(cell.coordinate)): return True
    return False
    
def getmerge(cell):
    for r in sheet_ranges.merged_cells.ranges:
        if r.issuperset(CellRange(cell.coordinate)): return r
    # return None
    return CellRange(cell.coordinate)


SUITE=r"""<testsuite name="%(suitename)s" >
<node_order><![CDATA[1]]></node_order>
<details><![CDATA[]]></details> 
%(testcases)s
</testsuite>
"""
HEADER = r"""<?xml version="1.0" encoding="UTF-8"?>
"""

TESTCASE = r"""<testcase internalid="198575" name="%(casename)s">
    <summary><![CDATA[<p>
    %(casename)s</p>
]]></summary>
    <preconditions><![CDATA[<p>
    %(preconditions)s</p>
]]></preconditions>
    <execution_type><![CDATA[1]]></execution_type>
    <importance><![CDATA[2]]></importance>
<steps>
<step>
    <step_number><![CDATA[1]]></step_number>
    <actions><![CDATA[<p>
    %(steps)s</p>
]]></actions>
    <expectedresults><![CDATA[<p>
    %(expect)s</p>
]]></expectedresults>
    <execution_type><![CDATA[1]]></execution_type>
</step>
</steps>
</testcase>
"""
    
    
for file in filename:
    print("\n"+file)
    result_dir = dir + file.replace(".xlsx","") + "\\"
    if not os.path.exists(result_dir):
        os.makedirs(result_dir)
        
    wb = load_workbook(file)

    print(wb.sheetnames)
    for sheetname in wb.sheetnames:
        print(sheetname)
        if sheetname == "Sheet1": continue

        sheet_ranges = wb[sheetname]

        suites = []

        i=11
        casename0=None
        preconditions0=""
        m0 = None
        pm0 = CellRange("A1:A1")
        cm0 = CellRange("A1:A1")
        suitename0 = ""
        suite = None
        j = 2
        cases = []
        while sheet_ranges['B'+str(i)].value:
            # print(sheet_ranges['B'+str(i)].value)
            #确定模块
            if sheet_ranges['C'+str(i)].value:
                if suitename0 != "":
                    print(suitename0+":"+str(len(cases)))
                    suite = SUITE % dict(suitename = suitename0,testcases = "".join(cases))
                    cases = []
                    suites.append(suite)
                m0 = getmerge(sheet_ranges['C'+str(i)])
                suitename0 = sheet_ranges['C'+str(i)].value
                suitename0 = suitename0.replace("\"","“").replace("&","_")
                # suite = SUITE% dict(suitename = suitename0)
            elif suitename0 == "" or not ismerged(sheet_ranges['C'+str(i)]) or getmerge(sheet_ranges['C'+str(i)]) != m0:
                if suitename0 != "":
                    print(suitename0+":"+str(len(cases)))
                    suite = SUITE % dict(suitename = suitename0,testcases = "".join(cases))
                    cases = []
                    suites.append(suite)
                m0 = getmerge(sheet_ranges['C'+str(i)])
                suitename0 = "未知模块"
                # suite = SUITE% dict(suitename = suitename0)
                
            #确定用例名
            #单元格不空，则以此为用例名，并且记录该单元格的范围
            if sheet_ranges['D'+str(i)].value:
                casename = casename0 = sheet_ranges['D'+str(i)].value.replace("&","_").replace("\"","“")
                cm0 = getmerge(sheet_ranges['D'+str(i)])
                j = 2
            #单元格是空的，但跟前面的单元格是同一范围，则用前面的用例名，加上序号
            elif getmerge(sheet_ranges['D'+str(i)]) == cm0:
                casename = "%s(%i)"%(casename0,j)
                j+=1
            #单元格是空的，跟前面的单元格不是同一范围，则取前置条件做用例名
            elif sheet_ranges['E'+str(i)].value:
                casename = sheet_ranges['E'+str(i)].value.replace("&","_").replace("\"","“")
            #前置条件也是空的，则取步骤为用例名
            elif sheet_ranges['F'+str(i)].value:
                casename = sheet_ranges['F'+str(i)].value.replace("&","_").replace("\"","“")
            #步骤也是空的，则取期望结果为用例名
            elif sheet_ranges['G'+str(i)].value:
                casename = sheet_ranges['G'+str(i)].value.replace("&","_").replace("\"","“")
            else:
                casename = "空行"
                
            # if sheet_ranges['D'+str(i)].value:
                # casename = sheet_ranges['D'+str(i)].value
                # casename = casename.replace("&","_").replace("\"","“")
                # casename0=casename
                # j = 2
            # elif casename0:
                # casename = "%s(%i)"%(casename0,j)
                # j+=1
            # elif sheet_ranges['E'+str(i)].value:
                # casename = sheet_ranges['E'+str(i)].value
            # else:
                # casename = sheet_ranges['F'+str(i)].value
                
            #确定前置条件
            #单元格不空，则以此为前置，并且记录该单元格的范围
            if sheet_ranges['E'+str(i)].value:
                preconditions = preconditions0 = sheet_ranges['E'+str(i)].value
                pm0 = getmerge(sheet_ranges['E'+str(i)])
            #单元格是空的，但跟前面的单元格是同一范围，则用前面的前置
            elif getmerge(sheet_ranges['E'+str(i)]) == pm0:
                preconditions = preconditions0
            #单元格是空的，并且跟前面的单元格不是同一范围，则前置为空，并且记录该单元格范围
            else:
                preconditions = preconditions0 = ""
                pm0 = getmerge(sheet_ranges['E'+str(i)])
            
            # if sheet_ranges['E'+str(i)].value:
                # preconditions = sheet_ranges['E'+str(i)].value
                # if ismerged(sheet_ranges['E'+str(i)]):
                    # preconditions0 = sheet_ranges['E'+str(i)].value
            # elif ismerged(sheet_ranges['E'+str(i)]):
                # preconditions=preconditions0
            # else:
                # preconditions = preconditions0 = ""
                
            #步骤和期望结果
            steps = sheet_ranges['F'+str(i)].value
            expect = sheet_ranges['G'+str(i)].value
            # print(sheet_ranges['B'+str(i)].value)
            # print(suitename0)
            # print("casename:"+str(casename))
            # print("preconditions:"+str(preconditions))
            # print("steps:"+str(steps))
            # print("expect:"+str(expect))
            
            case = TESTCASE % dict(
                            casename=casename,
                            preconditions=preconditions,
                            steps=steps,
                            expect=expect,
            )
            
            # print(case)
            cases.append(case)
            i += 1
        print(suitename0+":"+str(len(cases)))
        suite = SUITE % dict(suitename = suitename0,testcases = "".join(cases))
        cases = []
        suites.append(suite)
            
        # print("".join(suites))
            
        final = SUITE % dict(suitename = sheetname.replace("&","_").replace("\"","“"),testcases = "".join(suites))
            
        f = open(result_dir+sheetname+".xml","wb")
        f.write(HEADER.encode("utf8"))
        f.write(final.encode("utf8"))
        f.close



