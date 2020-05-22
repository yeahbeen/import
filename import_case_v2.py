'''
作用：
    把excel的用例转化成能导入到testlink的xml
使用方法：
    1、把用例Excel文件和该脚本放到同一个目录
    2、命令行进入该目录，执行python import_case.py
    3、执行结束后，就会生成能导入到testlink的xml，每个Excel对应一个文件夹，Excel的每个sheet对应一个xml。
使用条件：
    1、安装Python3，安装openpyxl包
    2、Excel要符合一定的格式：
        1）、第一列留空，从第二列(B列)开始，依次是：用例编号、模块1、模块2...用例名称、前置条件、操作步骤、期望输出    
        2）、正式的用例从11行开始，11行以前的内容忽略
        3）、每个用例都要有编号，编号为空将会结束解析
        4）、Sheet1会被忽略，后面每个sheet作为一个大模块
注意：
    1、用例名称不能为空，而且应该是在一个单元格里面，不要在合并单元格里面
'''
import os,sys
from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange
import xml.etree.ElementTree as ET

dir = __file__+"\\..\\"

filename = []
for file in os.listdir(dir):
    # print(file)
    # print(os.path.splitext(file)[1])
    if os.path.splitext(file)[1] == ".xlsx": 
        filename.append(file)
        
if filename == []:
    print("没有Excel文件")
    sys.exit()
    
print(filename)
        
def ismerged(cell):
    for r in sheet_ranges.merged_cells.ranges:
        if r.issuperset(CellRange(cell.coordinate)): return True
    return False
    
def getmerge(cell):
    for r in sheet_ranges.merged_cells.ranges:
        if r.issuperset(CellRange(cell.coordinate)): return r
    # return None
    return CellRange(cell.coordinate)
    
    
    
for file in filename:
    print("\n"+file)
    result_dir = dir + file.replace(".xlsx","") + "\\"
    if not os.path.exists(result_dir):
        os.makedirs(result_dir)
        
    wb = load_workbook(file)

    print(wb.sheetnames)
    for sheetname in wb.sheetnames:
        print(sheetname)
        if sheetname == "Sheet1": continue

        sheet_ranges = wb[sheetname]

        # i=11
        
        i = 0
        for r in range(1,20):  # 找出从第几行开始
            print(sheet_ranges["B"+str(r)].value)
            if sheet_ranges["B"+str(r)].value == "用例编号":
                i = r
                break
        if i == 0:
            print("请确保包含“用例编号”!")
            os.exit()
        print("标题所在行号:"+str(i))
        i = i+1
        
        fi = 0
        for ii in range(20):  # 找出有几个模块
            # print(sheet_ranges[chr(66+ii)+"11"].value)
            if sheet_ranges[chr(66+ii)+str(i-1)].value == "用例名称":
                fi = ii
                break
        if fi == 0:
            print("请确保有一列是“用例名称”!")
            os.exit()
        print("用例名称列号:"+str(fi+2))
        
        root = ET.Element("testsuite",{"name":sheetname})

        while sheet_ranges['B'+str(i)].value:
            # print(sheet_ranges['B'+str(i)].value)
            for iii in range(fi):
                xpath = "."
                pxpath = ""
                for iiii in range(iii):
                    # print("iiii:"+str(iiii)+"   iii:"+str(iii))
                    rangeco = sheet_ranges[getmerge(sheet_ranges[chr(67+iiii)+str(i)]).coord]
                    if type(rangeco) == tuple:
                        mdname = rangeco[0][0].value
                    else:
                        mdname = rangeco.value
                    # print(mdname)                
                    if mdname is None:
                        # mdname = "子模块"
                        continue
                    if iiii == iii - 1:
                        pxpath = xpath
                    xpath +="/testsuite[@name='%s']"%(mdname)
                # print("xpath:"+xpath)
                # print("pxpath:"+pxpath)
                if root.findall(xpath) == []:
                    tmp_suit = ET.SubElement(root.findall(pxpath)[0],"testsuite",{"name":mdname})
            tmp_suit = root.findall(xpath)[0]
            casename = sheet_ranges[chr(66+fi)+str(i)].value
            if casename is None:  #前置条件为空，如果用例名不空且不是合并单元格，则取用例名，否则取操作步骤
                if sheet_ranges[chr(66+fi+1)+str(i)].value is not None and not ismerged(sheet_ranges[chr(66+fi+1)+str(i)]):
                    casename = sheet_ranges[chr(66+fi+1)+str(i)].value
                else:
                    casename = sheet_ranges[chr(66+fi+2)+str(i)].value
            print("casename:"+casename)
            testcase = ET.SubElement(tmp_suit,"testcase",{"name":casename})
            summary = ET.SubElement(testcase,"summary")
            summary.text = r"<![CDATA[<p>%s</p>]]>"%(casename)
            # preconditions = ET.SubElement(testcase,"preconditions")
            # preconditions.text = r"<![CDATA[<p>%s</p>]]>"%(casename)
            rangepre = sheet_ranges[getmerge(sheet_ranges[chr(66+fi+1)+str(i)]).coord]
            if type(rangepre) == tuple:
                precod = rangepre[0][0].value
            else:
                precod = rangepre.value
            print("precondition:"+str(precod))
            preconditions = ET.SubElement(testcase,"preconditions")
            preconditions.text = r"<![CDATA[<p>%s</p>]]>"%(str(precod))
            execution_type = ET.SubElement(testcase,"execution_type")
            execution_type.text = r"<![CDATA[1]]>"
            importance = ET.SubElement(testcase,"importance")
            importance.text = r"<![CDATA[2]]>"
            steps = ET.SubElement(testcase,"steps")
            step = ET.SubElement(steps,"step")
            step_number = ET.SubElement(step,"step_number")
            step_number.text = r"<![CDATA[1]]>"
            actions = ET.SubElement(step,"actions")
            actions.text = r"<![CDATA[<p>%s</p>]]>"%(sheet_ranges[chr(66+fi+2)+str(i)].value)
            expectedresults = ET.SubElement(step,"expectedresults")
            expectedresults.text = r"<![CDATA[<p>%s</p>]]>"%(sheet_ranges[chr(66+fi+3)+str(i)].value)
            execution_type2 = ET.SubElement(step,"execution_type")
            execution_type2.text = r"<![CDATA[1]]>"
            i += 1
        # ET.dump(root)
        tree = ET.ElementTree(root)
        tree.write("temp.xml",encoding="utf8")        
        with open("temp.xml","r",encoding='UTF-8') as f:
            treestr = f.read()
        os.remove("temp.xml")
        treestr = treestr.replace('&lt;','<').replace('&gt;','>')
        print(treestr)
        with open(result_dir+sheetname+".xml","w",encoding='UTF-8') as f:
            f.write(treestr)


